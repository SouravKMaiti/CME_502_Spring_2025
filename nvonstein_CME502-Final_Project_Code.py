#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 15 16:22:44 2025

@author: nvonstein

NOTE: Please see "Main" at the bottom of the file for details on how the script is meant to operate.
"""

import os
import pandas as pd
# pip install pandas openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from PIL import Image
from scipy import stats
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score, root_mean_squared_error


def verify_folder_setup(folder):
    # Purpose: To verify that the folder contains the correct calibration file, scan number folders
    # Requirement: The calibration standards folder needs to have "Calibration Standards.xlsx" in the name and must be
    # formatted as shown in the word document outlining this program.

    # 1. Verify the calibration file is in the folder
    contents = [os.path.join(folder, item) for item in os.listdir(folder)]
    calibration_file = ""
    calibration_file_found = False
    for content in contents:
        if "Calibration Standards.xlsx" in content:
            calibration_file = content
            calibration_file_found = True
    if not calibration_file_found:
        print("The calibration file was not found.")
        print("Note: the calibration file must end with: Calibration Standards.xlsx")
        return
    else:
        print("Calibration file found!")
        print(calibration_file)

    print()

    # 2. Create the dataframe and confirm all of the scan number folders are present
    df = pd.read_excel(calibration_file)
    scan_folders = []
    scan_numbers = []
    all_folders = all_folders = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isdir(os.path.join(folder, f))]
    all_scan_folders_found = True
    for scan in df["Scan #"]:
        found_scan_number_folder = False
        for folder in all_folders:
            if str(scan) in folder:
                found_scan_number_folder = True
                scan_folders.append(folder)
                scan_numbers.append(str(scan))
                break
        if not found_scan_number_folder:
            all_scan_folders_found = False
            print(f"The folder for Scan Number: {scan} was not found!")
    if not all_scan_folders_found:
        print("Cannot Continue. Not all scan folders found.")
        return
    else:
        print("All scan folders found!")

    #scan_numbers.sort()
    #scan_folders.sort()
    return df, calibration_file, scan_numbers, scan_folders


def get_elements(df):
    elements = []
    headers = df.columns.tolist()
    for header in headers:
        if "Intensity" in header:
            elements.append(header[:header.index("Intensity")-1])
    return elements

def get_x_data(txt_file):
    with open(txt_file,"r") as file:
        return file.readline().strip().split()

def get_y_data(txt_file):
    with open(txt_file,"r") as file:
        rows = file.readlines()
        y_data = []
        for row in rows:
            y_data.append(row.split()[0])
    return y_data

def get_intensity_data(txt_file):
    data = []
    with open(txt_file, "r") as file:
        for line in file:
            values = line.split()
            row = []
            for value in values:
                row.append(float(value))
            data.append(row)
    return data

def add_data_to_excel_sheet(excel_file,x_data,y_data,intensity_data):
    wb = load_workbook(excel_file)
    wb.create_sheet(scan_number)
    ws = wb[scan_number]
    x_data.insert(0,"1")
    x_data = [float(value) for value in x_data]

    for col, value in enumerate(x_data,start=1):
        ws.cell(row=1,column=col,value=value)
    for row, value in enumerate(y_data,start=2):
        ws.cell(row=row,column=1,value=value)

    for i, row in enumerate(intensity_data,start=2):
        for j, value in enumerate(row,start=2):
            ws.cell(row=i,column=j,value=value)

    wb.save(excel_file)

def normalize_to_zero(data):
    if not data:
        raise ValueError("Input list is empty.")

    # Flatten if input is list of [x]
    flat_data = [item[0] if isinstance(item, list) else item for item in data]

    # Convert to float
    float_data = [float(x) for x in flat_data]
    base = float_data[0]

    # Normalize and round to 2 decimal places, then convert to strings
    normalized = [str(round(x - base, 2)) for x in float_data]
    return normalized


def average_from_list(data):
    if not data or not any(data):
        raise ValueError("Input data is empty or invalid.")

    total = 0.0
    count = 0

    for row in data:
        for item in row:
            total += float(item)
            count += 1

    if count == 0:
        raise ValueError("No valid float values found.")


    return total / count

def process_scan(folder,heatmap_folder,excel_file,scan_folder,scan_number,element,vmin=0,vmax=0.5):

    SRX_files = [os.path.join(scan_folder, item) for item in os.listdir(scan_folder)]
    x_found, y_found, intensity_found, not_run = False, False, False, True
    heatmap_paths = []
    average_intensity_data = ""
    for SRX_file in SRX_files:

        if "x_pos.txt" in SRX_file:
            x_data = get_x_data(SRX_file)
            x_data = normalize_to_zero(x_data)
            x_found = True
        if "y_pos.txt" in SRX_file:
            y_data = get_y_data(SRX_file)
            y_data = normalize_to_zero(y_data)
            y_found = True
        if "detsum_" + element + "_K_norm.txt" in SRX_file:
            intensity_data = get_intensity_data(SRX_file)
            average_intensity_data = average_from_list(intensity_data)
            intensity_found = True
        if x_found and y_found and intensity_found and not_run:
            add_data_to_excel_sheet(excel_file,x_data,y_data,intensity_data)
            title = element+"-"+str(scan_number)+".png"
            heatmap_paths = save_heatmap(x_data[1:], y_data, intensity_data, heatmap_folder, title,title=title,vmin=vmin,vmax=vmax)
            not_run = False
    return heatmap_paths, average_intensity_data

def create_subfolder(parent_folder, subfolder_name):
    new_folder_path = os.path.join(parent_folder, subfolder_name)

    # Create the subfolder if it doesn't already exist
    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)
        print(f"Created subfolder: {new_folder_path}")
    else:
        print(f"Subfolder already exists: {new_folder_path}")

    return new_folder_path

def save_heatmap(x, y, data, output_folder, filename,
                 title="Heatmap", xlabel="pixels", ylabel="pixels",
                 cmap='jet', dpi=100,
                 vmin=0, vmax=0.2):

    os.makedirs(output_folder, exist_ok=True)
    data_array = np.array(data)

    # Cell size in inches
    cell_size = 0.25
    height = cell_size * len(data_array)
    width = cell_size * len(data_array[0])

    fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)

    sns.heatmap(data_array,
                xticklabels=x,
                yticklabels=y,
                cmap=cmap,
                cbar=True,
                square=True,
                ax=ax,
                vmin=vmin,
                vmax=vmax)

    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)

    output_path = os.path.join(output_folder, filename)
    plt.savefig(output_path, dpi=dpi, bbox_inches='tight', pad_inches=0.1)
    plt.close()
    return output_path

def combine_images_horizontally(image_paths, output_path):
    # Load all images
    images = [Image.open(path) for path in image_paths]

    # Get total width and max height
    total_width = sum(img.width for img in images)
    max_height = max(img.height for img in images)

    # Create blank canvas
    combined_img = Image.new('RGB', (total_width, max_height), color=(255, 255, 255))

    # Paste each image
    x_offset = 0
    for img in images:
        combined_img.paste(img, (x_offset, 0))
        x_offset += img.width

    # Save result
    combined_img.save(output_path)

def plot_model(ax_row, x, y, y_pred, residuals, title, color,slope,intercept):
    equation = f'$y = ({slope:.2e})x + ({intercept:.2e})$'
    xlabel = "Concentration"
    ylabel="Intensity"

    # Fit
    axs[ax_row, 0].scatter(x, y, label='Data', color="black")
    axs[ax_row, 0].plot(
        x, y_pred, '--', color=color,
        label=(
            f'{title}\n'
            f'{equation}\n'
            f'$R^2$={r2_score(y, y_pred):.4f}, '
            f'RMSE={root_mean_squared_error(y, y_pred):.4f}, '
            f'MAE={mean_absolute_error(y, y_pred):.4f}')
    )
    axs[ax_row, 0].set_title(f'{title} Fit')
    axs[ax_row, 0].set_xlabel(xlabel)
    axs[ax_row, 0].set_ylabel(ylabel)
    axs[ax_row, 0].legend(
        loc='lower center',
        bbox_to_anchor=(0.5, 1.15),
        ncol=1,
        fontsize='small',
        frameon=False
    )
    axs[ax_row, 0].grid(True)

    # Residuals
    axs[ax_row, 1].scatter(x, residuals, color=color)
    axs[ax_row, 1].axhline(0, color=color, linestyle='--')
    axs[ax_row, 1].set_title(f'Residuals ({title})')
    axs[ax_row, 1].set_xlabel(xlabel)
    axs[ax_row, 1].set_ylabel('Residuals')
    axs[ax_row, 1].grid(True)

    # Histogram
    axs[ax_row, 2].hist(residuals, bins=10, color=color, edgecolor=color)
    axs[ax_row, 2].set_title(f'Residual Histogram ({title})')
    axs[ax_row, 2].set_xlabel('Residuals')
    axs[ax_row, 2].set_ylabel('Frequency')
    axs[ax_row, 2].tick_params(axis='x', labelsize=10)

# MAIN - The code below follows through each step of the script written for SRX data workup.

# Define the folder with Raw Data
folder = "/Users/nvonstein/Library/CloudStorage/Box-Box/_NV-Classes/CME502-Modeling/Final Project/250430-Raw Data/"

# Verify the folder setup
df, calib_excel, scan_numbers, scan_folders = verify_folder_setup(folder)

# Retrieve the available elements from the files in each scan folder
elements = get_elements(df)

# Create the "Heatmaps" sub folder to store heatmap images
heatmap_folder = create_subfolder(folder, "heatmaps")

# Iterate through all elements and perform the following:
for element in elements:
    heatmap_paths = []
    average_intensity_data = []
    
    # Create a new Excel File for the element's SRX data to save
    excel_file = os.path.join(folder,  "Intensity_Data_"+element+".xlsx")
    wb = Workbook()
    wb.active
    wb.save(excel_file)
    # Iterate scan folder, find the element of interest, add scan information to the excel file,
    # create the heatmap, and save the file to the heatmaps folder.
    for scan_number in scan_numbers:
        for scan_folder in scan_folders:
            if scan_number in scan_folder:
                # Establish the intensity range coloring for the heatmap
                vmin = 0
                vmax = 0.005
                heatmap,average_intensity = process_scan(folder,heatmap_folder,excel_file,scan_folder,scan_number,element,vmin=vmin,vmax=vmax)
                average_intensity_data.append(average_intensity)
                heatmap_paths.append(heatmap)

    # Removes the "Sheet" from the beginning of the excel file
    wb = load_workbook(excel_file)
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(excel_file)
    
    # Creates one, large png for all element specific heatmaps for easy comparisons
    combine_images_horizontally(heatmap_paths, excel_file.replace(".xlsx", ".png"))

    # Add the average intensity data to the df:
    for i in range(0,len(average_intensity_data)):
        df.loc[i,element+" Intensity"] = float(average_intensity_data[i])

    # Plot and perform a statistical analysis of the calibration curve
    if any(("Element "+element) in col for col in df.columns):
        x = df["Concentration of Element " +element]
        y = df[element+" Intensity"]

        print(x.shape,y.shape)

        # Linear Regression Model
        slope, intercept, r_value_lin, p_value, std_err = stats.linregress(x, y)
        y_pred_lin = intercept + slope * x
        residuals_lin = y - y_pred_lin
        rmse_lin = root_mean_squared_error(y, y_pred_lin)
        mae_lin = mean_absolute_error(y, y_pred_lin)
        r2_lin = r2_score(y, y_pred_lin)

        # Table comparing the different models:
        model = ['Linear Regression']
        rmses = [rmse_lin]
        maes = [mae_lin]
        r2s = [r2_lin]

        results_df = pd.DataFrame({
            'Model': model,
            'RMSE': rmses,
            'MAE': maes,
            'R²': r2s})
        print(results_df)

        fig, axs = plt.subplots(1, 3, figsize=(12, 3),squeeze=False)

        plot_model(0, x, y, y_pred_lin, residuals_lin, 'Linear Regression', "red",slope,intercept)
        # Adjust layout
        plt.tight_layout()
        plt.show()
        fig.savefig(calib_excel.replace("Standards.xlsx","Fit-"+element+".png"))

# Paste the dataframe back into the excel file:
wb = load_workbook(calib_excel)
ws = wb.active
for i, row in df.iterrows():
    for j, value in enumerate(row):
        column_name = df.columns[j]

        # Check the header and conditionally convert to string
        if column_name == "Scan #":
            ws.cell(row=2 + i, column=1 + j, value=str(int(value)))
        else:
            ws.cell(row=2 + i, column=1 + j, value=value)
# Save the changes
wb.save(calib_excel)

print("Finished")



